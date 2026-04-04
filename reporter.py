import yaml
from pathlib import Path
from openpyxl import Workbook, load_workbook
from database import FileDB


class Reporter:
    "Generate xlsx report on duplicate files and similar folders"
    def __init__(self, config_path):
        self.cfg = yaml.safe_load(Path(config_path).read_text())
        self.db = FileDB(self.cfg["database"]["path"])

    def _summary_data(self):
        rows = self.db.conn.execute(
            """SELECT ss.folder_root, ss.total_folders, ss.total_files, ss.scanned_folders, ss.scanned_files,
                      (SELECT COUNT(*) FROM similarity_results sr
                       JOIN folders da ON sr.folder_a_id = da.id
                       WHERE da.path LIKE ss.folder_root || '%'),
                      (SELECT COUNT(*) FROM file_matches fm
                       JOIN files f ON fm.file_a_id = f.id
                       JOIN folders d ON f.folder_id = d.id
                       WHERE fm.full_hash_match = 1 AND d.path LIKE ss.folder_root || '%'),
                      (SELECT COALESCE(SUM(f.size), 0) / 1e9 FROM file_matches fm
                       JOIN files f ON fm.file_a_id = f.id
                       JOIN folders d ON f.folder_id = d.id
                       WHERE fm.full_hash_match = 1 AND d.path LIKE ss.folder_root || '%')
               FROM scan_stats ss""").fetchall()
        return rows

    def _duplicate_files_data(self):
        hdrs = ["file_a", "file_b", "filename", "size_bytes", "folder_a", "folder_b", "folder_similarity"]
        rows = self.db.conn.execute(
            """SELECT fa.path, fb.path, fa.filename, fa.size, da.path, db.path, sr.jaccard_score
               FROM file_matches fm
               JOIN similarity_results sr ON fm.result_id = sr.id
               JOIN files fa ON fm.file_a_id = fa.id
               JOIN files fb ON fm.file_b_id = fb.id
               JOIN folders da ON fa.folder_id = da.id
               JOIN folders db ON fb.folder_id = db.id
               WHERE fm.full_hash_match = 1
               ORDER BY fa.size DESC""").fetchall()
        return hdrs, rows

    def _similar_folders_data(self):
        hdrs = ["folder_a", "folder_b", "jaccard", "shared_sigs", "files_a", "files_b", "bytes_a", "bytes_b", "confirmed_dups", "dup_bytes"]
        rows = self.db.conn.execute(
            """SELECT da.path, db.path, sr.jaccard_score, sr.shared_count,
                      da.file_count, db.file_count, da.total_bytes, db.total_bytes,
                      (SELECT COUNT(*) FROM file_matches fm WHERE fm.result_id = sr.id AND fm.full_hash_match = 1),
                      (SELECT SUM(fa.size) FROM file_matches fm JOIN files fa ON fm.file_a_id = fa.id
                       WHERE fm.result_id = sr.id AND fm.full_hash_match = 1)
               FROM similarity_results sr
               JOIN folders da ON sr.folder_a_id = da.id
               JOIN folders db ON sr.folder_b_id = db.id
               ORDER BY 10 DESC NULLS LAST""").fetchall()
        return hdrs, rows

    def _write_sheet(self, ws, hdrs, rows):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row): 
            for cell in row: cell.value = None
        ws.delete_rows(1, ws.max_row)
        ws.append(hdrs)
        for r in rows: ws.append(list(r))

    def _get_or_create_sheet(self, wb, name):
        if name in wb.sheetnames: return wb[name]
        return wb.create_sheet(name)

    def run(self, out_path=Path("filescan_report.xlsx")):
        if out_path.exists():
            wb = load_workbook(out_path)
        else:
            wb = Workbook()
            wb.active.title = "Summary"

        sum_hdrs = ["Folder", "Total Folders", "Total Files", "Scanned Folders", "Scanned Files",
                     "Similar Folder Pairs", "Duplicate Files", "Duplicate GB"]
        sum_rows = self._summary_data()
        ws_sum = self._get_or_create_sheet(wb, "Summary")
        self._write_sheet(ws_sum, sum_hdrs, sum_rows)

        hdrs, rows = self._duplicate_files_data()
        ws_dup = self._get_or_create_sheet(wb, "Duplicate Files")
        self._write_sheet(ws_dup, hdrs, rows)
        total_dup_bytes = sum(r[3] for r in rows)

        hdrs, rows = self._similar_folders_data()
        ws_sim = self._get_or_create_sheet(wb, "Similar Folders")
        self._write_sheet(ws_sim, hdrs, rows)

        wb.save(out_path)
        print(f"Summary: {len(sum_rows)} folder roots")
        print(f"Duplicate files: {total_dup_bytes/1e9:.2f} GB reclaimable")
        print(f"Written to {out_path.resolve()}")
        self.db.close()
